import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
import openpyxl
from openpyxl.styles import PatternFill, Font

def scrape_books(pages=1):
    all_books = []

    for page_num in range(1, pages + 1):
        print(f"Processing page {page_num} of {pages}...")

        url = f"http://books.toscrape.com/catalogue/page-{page_num}.html"

        response = requests.get(url)
        if (response.status_code != 200):
            print(f"Could not load page {page_num}. Status code: {response.status_code}")
            break

        soup = BeautifulSoup(response.content, "html.parser")
        books = soup.find_all("article", class_="product_pod")

        for book in books:
            full_title = book.h3.a["title"]
            price_text = book.find("p", class_="price_color").text
            price = float(price_text.replace("£", "").strip())

            all_books.append({
                "Title": full_title,
                "Price (£)": price,
                "Page": page_num
            })

        time.sleep(1)
    
    return all_books

def save_as_table(all_books):
    df = pd.DataFrame(all_books)

    output_filename = "results.xlsx"
    df.to_excel(output_filename, index=False)

def style_excel(filename="results.xlsx", price_threshold=20.00):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    dark_green_font = Font(color="006100", bold=True)

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
    
    for row in range(2, ws.max_row + 1):
        price_cell = ws[f'B{row}']

        if (price_cell.value and float(price_cell.value) < price_threshold):
            price_cell.fill = green_fill
            price_cell.font = dark_green_font

    for col in ws.columns:
        max_len = max(len(str(cell.value or ' ')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 8)

    wb.save(filename)


if __name__ == "__main__":
    pages_to_scrape = 5
    price_threshold = 20.00

    books_data = scrape_books(pages=pages_to_scrape)
    save_as_table(books_data)
    style_excel("results.xlsx", price_threshold=price_threshold)
    
    print(f"Scraping completed. Data saved to 'results.xlsx'.")